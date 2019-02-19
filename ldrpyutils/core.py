import argparse
import json
import logging
import os
import os.path
import sys
import re

try:
    from urllib.parse import urlparse
except ImportError:
     from urlparse import urlparse

import pkg_resources
import rdflib
import requests
import validators
from openpyxl import load_workbook
from rdflib import RDF, URIRef, Literal
from rdflib.namespace import NamespaceManager, Namespace

DATA_PATH = pkg_resources.resource_filename("ldrpyutils", 'data/')
logging.basicConfig()

def load_simple_file(excel_file,  user=None, passwd=None, emitFile=False, registry_auth_url=None,
                     updateOnlineRegisters=False,
                     verbose=False):
    wb = load_workbook(excel_file, data_only=True) #last computed value

    if verbose:
        print(wb.get_sheet_names())

    sheetsDict = {}
    reginfo_ws = False
    for sheet in wb:
        sheetname = sheet.title
        if sheetname == 'registerinfo':
            reginfo_ws = sheet

        else:
            regitems_ws = sheet
            sheetsDict[sheetname] = regitems_ws

    #process the register info
    reginfo_obj = get_registerinfo(reginfo_ws)

    #process the register items
    regitems_obj = process_all_registeritems_in_dict(sheetsDict)

    if verbose:
        print(reginfo_obj)
        print(regitems_obj)


    (g, status) = build_graph_and_post(reginfo_obj, regitems_obj,
                        user=user, passwd=passwd,
                        emitFile=emitFile,
                        registry_auth_url=registry_auth_url,
                        updateOnlineRegisters=updateOnlineRegisters,
                        verbose=verbose
                        )
    if verbose:
        print(status)
    return status['isSuccessful']

def load_multi_register_file(excel_file, user=None, passwd=None, emitFile=False, registry_auth_url=None,
                             updateOnlineRegisters=False,
                             verbose=False):
    wb = load_workbook(excel_file, data_only=True) #last computed value
    if verbose:
        print(wb.get_sheet_names())

    sheetsDict = {}
    reginfo_ws = False
    for sheet in wb:
        sheetname = sheet.title
        if sheetname == 'registerinfo':
            reginfo_ws = sheet

        else:
            regitems_ws = sheet
            sheetsDict[sheetname] = regitems_ws

    #process the register info
    reginfo_obj = get_registerinfo_multi_register(reginfo_ws)

    #process the register items
    regitems_obj = process_all_registeritems_in_dict(sheetsDict, verbose)

    if verbose:
        print(reginfo_obj)
        print(regitems_obj)


    #build the graph
    (g, status) = build_graph_and_post(reginfo_obj, regitems_obj, user=user, passwd=passwd, mode='multi',
                        emitFile=emitFile, registry_auth_url=registry_auth_url,
                        updateOnlineRegisters=updateOnlineRegisters,
                        verbose=verbose)
    if verbose:
        print(status)
    return status['isSuccessful']

def get_registerinfo(ws):
    if ws == False:
        return False

    registerinfo_obj = parse_sheet(ws)

    result = {}
    for key, value in registerinfo_obj['items']:
        result[key] = value

    return result



def process_all_registeritems_in_dict(sheetsDict, verbose=False):
    if sheetsDict == False:
        return False

    regItemsObj = {}
    for registerid, ws in sheetsDict.items():
        res = get_registeritems(registerid, ws, verbose=verbose)
        arrResults = []
        for i, rowvalues in enumerate(res['items']):
            result = {}

            for j, cellvalues in enumerate(rowvalues):
               key = res['header'][j]
               result[key] = cellvalues
            arrResults.append(result)
        regItemsObj[registerid] = arrResults

    return regItemsObj

def get_registeritems(registerid, sheet, verbose=False):
    if verbose:
        print("reg id: " + registerid)
    return parse_sheet(sheet)


def parse_sheet(sheet):
    # parse header
    first_row = sheet[1]
    header = []
    for cell in first_row:
        #print cell.value
        header.append(cell.value)

    count = 0
    arrItems = []
    for row in sheet.iter_rows():
        count = count + 1

        if count > 1:
            currArr = []
            for cell in row:
                #print cell.value
                currArr.append(cell.value)

            arrItems.append(currArr)

    result = {}
    result['header'] = header
    result['items'] = arrItems
    return result

def resource_path(relative):
        if hasattr(sys, "_MEIPASS"):
            return os.path.join(sys._MEIPASS, relative)
        return os.path.join(relative)


def build_graph_and_post(reginfo_obj, regitems_obj,
                user=None, passwd=None, mode='single', emitFile=False,
                registry_auth_url=None,
                updateOnlineRegisters=False,
                verbose=False):
    if reginfo_obj == False or regitems_obj == False :
        return False


    ns_prefix_lookup = {
        "description" : 'dct',
        "source": 'dct',
        "definition": 'skos',
        "broader": 'skos',
        "notation": 'reg',
        "note": 'skos',
        "altLabel": 'skos',
        "hiddenLabel": 'skos',
        "exactMatch": 'skos',
        "label" : 'rdfs',
    }

    prefixes_g = rdflib.Graph()
    if verbose:
        print("Prefix file...")
        print(__file__)
    PREFIX_FILE = pkg_resources.resource_filename("ldrpyutils", 'data/prefixes.ttl')
    #if(pkg_resources.resource_exists("ldrpyutils", 'data/prefixes.ttl')):
    #    if verbose:
    #        print("Prefix file exists")
    #        print(pkg_resources.resource_string("ldrpyutils", 'data/prefixes.ttl'))
    #else:
    #    if verbose:
    #        print("Prefix file does not exist!")

    if verbose:
        print(PREFIX_FILE)
    with open(PREFIX_FILE) as f:
        #read_data = f.read()
        prefixes_g.parse(f, format="ttl")
    nsMgr = NamespaceManager(prefixes_g)

    all_ns = [n for n in nsMgr.namespaces()]
    prefix_idx = {}
    for prefix, namespace in all_ns:
        #print (prefix, namespace.n3())
        prefix_idx[prefix] = Namespace(namespace)


    g = None
    status = {
        "didEmitFile" : False,
        "didUpdateOnlineRegisters": False,
        "isSuccessful": False,
    }
    if mode == 'single':
        register_id = reginfo_obj['id']
        register_url = reginfo_obj['registry_location']
        reglabel = reginfo_obj['label']
        regdescription = reginfo_obj['description']
        register_url = reginfo_obj['registry_location']

        (parent_reg_url, sub_reg_id) = get_register_location_parent_and_subreg_url(register_url)
        subreg_graph = get_subregister_graph(sub_reg_id, reglabel, regdescription, prefix_idx, nsMgr)
        subreg_data = subreg_graph.serialize(None, format='turtle')
        if verbose:
            print("Outputting register graph for " + sub_reg_id)
            print(subreg_data)
        g = get_register_graph(sub_reg_id, reginfo_obj, regitems_obj[sub_reg_id], nsMgr, prefix_idx, ns_prefix_lookup)
        data = g.serialize(None,format='turtle')
        if verbose:
            print("Outputting graph for " + sub_reg_id)
            print(data)
        if emitFile or updateOnlineRegisters:
            filename = sub_reg_id + ".ttl"
            g.serialize(filename, format="turtle")
            status['didEmitFile'] = True
            if updateOnlineRegisters:
                # use the file to update the registers
                resFlag = post_update_to_online_register(sub_reg_id, parent_reg_url, register_url, data, subreg_data,
                                               registry_auth_url=registry_auth_url,
                                               user=user, passwd=passwd,
                                               verbose=verbose
                                               )
                status['didUpdateOnlineRegisters'] = resFlag
                if resFlag == False:
                    status['isSuccessful'] = False
                else:
                    status['isSuccessful'] = True

    else:
        #assume multi register
        for key in reginfo_obj:
            register_id = key
            register_url = reginfo_obj[key]['registry_location']
            reglabel = reginfo_obj[key]['label']
            regdescription = reginfo_obj[key]['description']
            register_url = reginfo_obj[key]['registry_location']

            (parent_reg_url, sub_reg_id) = get_register_location_parent_and_subreg_url(register_url)
            subreg_graph = get_subregister_graph(sub_reg_id, reglabel, regdescription, prefix_idx, nsMgr)
            subreg_data = subreg_graph.serialize(None, format='turtle')


            g = get_register_graph(sub_reg_id, reginfo_obj[key], regitems_obj[key], nsMgr, prefix_idx, ns_prefix_lookup)
            data = g.serialize(format='turtle')
            status['didEmitFile'] = True
            if verbose:
                print(data)
            if emitFile:
                filename = sub_reg_id + ".ttl"
                g.serialize(filename, format="turtle")
            if updateOnlineRegisters:
               #use the file to update the registers
               resFlag = post_update_to_online_register(sub_reg_id, parent_reg_url, register_url, data, subreg_data,
                                                   registry_auth_url=registry_auth_url,
                                                   user=user, passwd=passwd,
                                                    verbose=verbose
                                                   )
               status['didUpdateOnlineRegisters'] = resFlag
               if resFlag == False:
                  status['isSuccessful'] = False
               else:
                  status['isSuccessful'] = True

    return (g, status)


def get_register_location_parent_and_subreg_url(register_url):
    #treat register_url as the final register i.e. subregister
    parsed = urlparse(register_url)
    path = parsed.path
    (head, tail) = os.path.split(path)
    head_url = parsed.scheme + "://" + parsed.netloc + head
    return (head_url, tail)


def get_subregister_graph(regid, reglabel, regdescription, prefix_idx, nsMgr):
    DCT = prefix_idx['dct']
    RDFS = prefix_idx['rdfs']
    RDF = prefix_idx['rdf']
    REG = prefix_idx['reg']
    subreg = None
    graph = rdflib.Graph(namespace_manager=nsMgr)
    try:
        subreg = URIRef(str(regid))
    except UnicodeEncodeError:
        subreg = URIRef(regid.encode('utf-8'))
    graph.add((subreg, RDF.type, REG.Register))
    graph.add((subreg, RDFS.label, Literal(str(reglabel))))
    graph.add((subreg, DCT.description, Literal(str(regdescription))))
    return graph
    """
    <register1>       a reg:Register ;
        rdfs:label       "Register 1"^^xsd:string ;
        dct:description  "Register containing a set of example concepts"^^xsd:string ;
        rdfs:member      reg1:conceptB , reg1:conceptC , reg1:conceptD , reg1:element-conf-report ;
.
    :return: 
    """


def processMultilineCell(data):
    if data is None:
        return None

    print("Applying regex")
    arr = re.split(r'[\n\r]+',data)
    print(arr)
    return arr

        #return str.splitlines(data)


def get_register_graph(register_id, register_info, register_items, nsMgr, prefix_idx, ns_prefix_lookup):
    DCT = prefix_idx['dct']
    SKOS = prefix_idx['skos']
    REG = prefix_idx['reg']
    RDFS = prefix_idx['rdfs']


    g = rdflib.Graph(namespace_manager=nsMgr)


    # process items
    dictConcepts = {}
    items_data = register_items
    try:

        for item in items_data:
            concept = None
            if item['id'] not in dictConcepts:
                if(item['id'] is None):
                    #check if this is blank row
                    print("id column is empty. checking if row is empty.")
                    if 'label' in item and item['label'] is None:
                        #confirmed that this is none, so skip this item
                        print("row is empty. skipping...")
                        continue
                    else:
                        print("row is not empty. attempting to create concept...")
                concept = create_concept_with_id(item['id'], g, prefix_idx)
                dictConcepts[item['id']] = concept
            else:
                concept = dictConcepts[item['id']]

            listSkosLexicalProps = ['altLabel', 'hiddenLabel', ]
            listSkosRelProps = ['broader', 'related', 'semanticRelation',
                                'mappingRelation', 'closeMatch', 'exactMatch', 'broadMatch', 'narrowMatch', 'relatedMatch']
            #iterate over the fields in the register
            for key in item:
                if key == None:
                    continue
                #and create kvp's for any header
                if key != 'id' and (key not in listSkosLexicalProps) and (key not in listSkosRelProps):
                    # get prefix
                    currPrefix = ns_prefix_lookup[key]
                    currNs = prefix_idx[currPrefix]
                    if item[key] != None:
                        g.add((concept, currNs[key], Literal(item[key])))

                #handle any special cases below

                # If this is a label field, register it as rdfs:label (default) and skos:prefLabel (addition)
                if key == 'label':
                    g.add((concept, SKOS.prefLabel, Literal(item[key])))

                if key == 'altLabel':
                    arrValues = processMultilineCell(item[key])
                    if arrValues is not None:
                        for v in arrValues:
                            print("AltLabel: " + v)
                            g.add((concept, SKOS.altLabel, Literal(v)))

                if key == 'hiddenLabel':
                    arrValues = processMultilineCell(item[key])
                    if arrValues is not None:
                        for v in arrValues:
                            print("HiddenLabel: " + v)
                            g.add((concept, SKOS.hiddenLabel, Literal(v)))


                if key == 'description':
                    g.add((concept, SKOS.definition, Literal(item[key])))

                if key == 'notation':
                    g.add((concept, SKOS.notation, Literal(item[key])))

                if key == 'broader':
                    #find the refering concept
                    broaderConceptId = item[key]
                    if(broaderConceptId is None or broaderConceptId == ''):
                        continue

                    if broaderConceptId in dictConcepts:
                        broaderConcept = dictConcepts[broaderConceptId]
                        g.add((concept, SKOS.broader, broaderConcept))
                        g.add((broaderConcept, SKOS.narrower, concept))
                    elif validate_url(broaderConceptId):
                        broaderConcept = URIRef(broaderConceptId)
                        g.add((concept, SKOS.broader, broaderConcept))
                        g.add((broaderConcept, SKOS.narrower, concept))
                    else:
                        #create it
                        if broaderConceptId is None:
                            print("hello")
                        broaderConcept = create_concept_with_id(broaderConceptId, g, prefix_idx)
                        dictConcepts[broaderConceptId] = broaderConcept

                        # add relations
                        g.add((concept, SKOS.broader, broaderConcept))
                        g.add((broaderConcept, SKOS.narrower, concept))
                        # add placeholder label for the broader concept
                        g.add((broaderConcept, RDFS.label, Literal(str(broaderConceptId))))
                elif key in listSkosRelProps:
                    uri = item[key]
                    if(uri is None or uri == ''):
                        continue

                    currProp = rdflib.term.URIRef(SKOS[key])
                    if validate_url(uri):
                        currConcept = URIRef(uri)
                        g.add((concept, currProp, currConcept))

    except ValueError as err:
        print(err)
        raise

    return g

def validate_url(urlstring):
    curr = str(urlstring)
    if curr.startswith("http"):
        return validators.url(curr)
    return False

def create_concept_with_id(id, graph, prefix_idx):
    DCT = prefix_idx['dct']
    SKOS = prefix_idx['skos']
    REG = prefix_idx['reg']
    test_graph = rdflib.Graph()
    if id is None:
        raise ValueError("id cannot be empty")

    #trim the id
    cleaned_id = str(id).strip()
    #check if id has any special characters - '/', '\', ':'
    if re.search('/|:/\\\\| |,|;', cleaned_id):
        raise ValueError("Error in id ('" + cleaned_id + "'). id cannot have spaces or special characters like '/', ':', '\\'.")

    try:
        #check that id is able to form a valid URI
        concept = URIRef(str(cleaned_id))
        test_graph.add((concept, RDF.type, SKOS.Concept))

        #test that we can serialize the statement - if so, all good; else it means URI is broken
        test_graph.serialize(format='turtle')
    except UnicodeEncodeError:
        concept = URIRef(cleaned_id.encode('utf-8'))
    except Exception:
        #assume we could not serialise the format
        raise ValueError("Error in trying to form a URI with '%s'." % cleaned_id)
    graph.add((concept, RDF.type, SKOS.Concept))
    return concept

def post_update_to_online_register(sub_reg_id, parent_reg_url, register_url, data, subreg_data, registry_auth_url=None, user=None, passwd=None,
                                   verbose=False, create_register_if_not_exists=True):
    s = requests.Session()
    r = s.post(registry_auth_url, data={'userid': user, 'password': passwd})

    resFlag = False

    if verbose:
        print(r.status_code)
        print(r.text)
        print(r.cookies.get_dict())

    if r.status_code == 200:
        #successfully logged in - so upload the data
        url = register_url
        headers = {"Content-Type": "text/turtle"}
        if verbose:
            print(s.cookies.get_dict())

        if create_register_if_not_exists:
            #create the register if not exists
            print("Creating register " + sub_reg_id + " in " + parent_reg_url)
            r = s.post(parent_reg_url, data=subreg_data, headers=headers)
            if verbose:
                print(r.status_code)
                #print(json.dumps(r.json))

        r=s.post(url, data=data, headers=headers)
        if verbose:
            print(r.status_code)
        if r.status_code == 201:
                resFlag = True
                print("Successfully created items in register '" + sub_reg_id + "'")
        elif r.status_code == 403:
            #force update
            url = register_url + "?edit"
            headers = {"Content-Type": "text/turtle"}
            if verbose:
                print(s.cookies.get_dict())

            r = s.post(url, data=data, headers=headers)

            if r.status_code == 204:
                resFlag = True
                print("Successfully updated register '" + sub_reg_id + "'")
            if verbose:
                print(r.status_code)
        if r.status_code == 400:
            print("Failed to post content to LDR.")
            print(r.reason)
            print(r.json)
    return resFlag


def get_registerinfo_multi_register(ws):
    if ws == False:
        return False

    registerinfo_obj = parse_sheet(ws)

    d = []
    for row in registerinfo_obj['items']:
        item = {}
        for i, elem in enumerate(row):
            item[registerinfo_obj['header'][i]] = elem
        d.append(item)

    result = {}
    for item in d:
        if item['Register_id'] in result:
            result[item['Register_id']][item['Register_property']] = item['Register_property_value']
        elif item['Register_id'] == None or item['Register_id'] == "":
            #do nothing
            abc = 123
        else:
           result[item['Register_id']] = {}
           result[item['Register_id']][item['Register_property']] = item['Register_property_value']
    return result

def excel2ldr():
    parser = argparse.ArgumentParser(description='Upload vocab content defined in Excel to a Linked Data Registry.')
    parser.add_argument('--user', action="store", dest="user", default=None, help="User name to authenticate with registry")
    parser.add_argument('--pass', action="store", dest="passwd", default=None, help="Password to authenticate with registry")
    parser.add_argument('--no-emitfiles', action="store_true", dest="noEmitFile", default=False, help="Flag to specify not to emit files")
    parser.add_argument('--multi', action="store_true", dest="isMulti", default=False, help="Flag to indicate whether the file contains multiple registers to update")
    parser.add_argument('--staging-only', action="store_true", dest="stagingOnly", default=False, help="Flag to stage outputs without pushing updates to registry")
    parser.add_argument('--registry-url', action="store", dest="registryUrl", default=None, help="Override and use this registry url instead")
    parser.add_argument('--verbose', action="store_true", dest="isVerbose", default=False, help="Flag to specify verbose output")

    parser.add_argument("excel_file", help="Path for the excel file")

    args = parser.parse_args()

    #user = 'bba-vocabs@csiro.au'
    #passwd = "abc123"

    #load defaults from config file
    cfg = None
    with open('config.json') as data_file:
        cfg = json.load(data_file)


    registry_url = cfg['registry_url']
    #check if user wants to override configuration registry url
    if args.registryUrl:
        registry_url = args.registryUrl

    registry_auth_url = registry_url + "/system/security/apilogin"


    #check if user wants no files emitted
    emitFile = True
    if args.noEmitFile:
        emitFile = False

    #check if user just wants to stage things
    updateOnlineRegisters = True
    if args.stagingOnly:
        updateOnlineRegisters = False

    verbose = False
    if args.isVerbose:
        verbose = True
        print("Verbose mode turned on")

    user = args.user
    passwd = args.passwd
    filepath = args.excel_file


    if args.isMulti:
        resultFlag = load_multi_register_file(filepath, user, passwd, emitFile=emitFile,
                                 registry_auth_url=registry_auth_url,
                                 updateOnlineRegisters=updateOnlineRegisters,
                                 verbose=verbose)

    else:
        resultFlag = load_simple_file(filepath, user, passwd, emitFile=emitFile, registry_auth_url=registry_auth_url,
                         updateOnlineRegisters=updateOnlineRegisters,
                         verbose=verbose
                         )
    


if __name__ == "__main__":
    excel2ldr()

